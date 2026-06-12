import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# (question, [a1,a2,a3,a4], correct_index_1based, time)
QUIZZES = {
"kahoot-1-navkar-core-sutras": ("Navkar Mantra & Core Sutras", [
("How many supreme beings does the Navkar Mantra salute?", ["Three","Five","Seven","Nine"], 2, 20),
("'Namo Arihantanam' bows to whom?", ["The liberated souls","Those who conquered inner enemies","The teachers","All monks"], 2, 20),
("'Namo Siddhanam' salutes whom?", ["Liberated souls in moksha","Spiritual teachers","Scripture scholars","All living beings"], 1, 20),
("Which sutra praises all 24 Tirthankaras by name?", ["Navkar Mantra","Logassa","Annath Sutra","Tikhutto"], 2, 20),
("Chattari Mangalam declares how many auspicious refuges?", ["Two","Three","Four","Five"], 3, 20),
("Which sutra repents harm to living beings while walking?", ["Iriyavahiyam","Tassa Uttari","Namutthunam","Jankinchi"], 1, 20),
("The Uvasaggaharam Stotra praises which Tirthankara?", ["Mahavir","Adinath","Parshvanath","Neminath"], 3, 20),
("Who composed the Uvasaggaharam Stotra?", ["Bhadrabahu Swami","Manatungasuri","Haribhadrasuri","Gautam Swami"], 1, 20),
("The Bhaktamar Stotra has how many verses (Shvetambar)?", ["24","44","48","108"], 3, 20),
("'Karemi Bhante' is the vow that begins which practice?", ["Pratikraman","Samayik","Chaityavandan","Paryushan"], 2, 20),
("Which sutra is recited before Kayotsarga meditation?", ["Annath Sutra","Logassa","Jankinchi","Navkar"], 1, 20),
("'Khamemi Savve Jiva' is a prayer of what?", ["Fasting","Universal forgiveness","Charity","Pilgrimage"], 2, 20),
]),
"kahoot-2-tirthankaras": ("The 24 Tirthankaras", [
("Who is the 1st Tirthankara?", ["Mahavir","Parshvanath","Rishabhanath (Adinath)","Shantinath"], 3, 20),
("Who is the 24th and last Tirthankara?", ["Neminath","Mahavir","Parshvanath","Mallinath"], 2, 20),
("What is Mahavir's symbol (lanchan)?", ["Elephant","Bull","Lion","Serpent"], 3, 20),
("What is Parshvanath's symbol?", ["Serpent","Conch","Lotus","Moon"], 1, 20),
("What is Adinath's symbol?", ["Horse","Bull","Deer","Lion"], 2, 20),
("Which Tirthankara has the Conch (Shankha) symbol?", ["Naminath","Munisuvrata","Neminath","Aranath"], 3, 20),
("Shantinath, the 16th Tirthankara, has which symbol?", ["Goat","Deer","Tortoise","Boar"], 2, 20),
("Where did Mahavir attain moksha?", ["Pavapuri","Girnar","Shatrunjay","Sammed Shikhar"], 1, 20),
("Mahavir was born in which kingdom/city?", ["Ayodhya","Kshatriyakund (Vaishali)","Varanasi","Rajgir"], 2, 20),
("How many years did Mahavir practice austerities before Kevalgyan?", ["7 years","12 years","30 years","40 years"], 2, 20),
("Which Tirthankara is worshipped as 'Chintamani' (wish-fulfilling jewel)?", ["Parshvanath","Padmaprabha","Chandraprabha","Vasupujya"], 1, 20),
("Chandraprabha, the 8th Tirthankara, has which symbol?", ["Sun","Moon","Star","Lotus"], 2, 20),
]),
"kahoot-3-samayik-pratikraman": ("Samayik & Pratikraman", [
("How long does one Samayik last?", ["24 minutes","48 minutes","60 minutes","12 minutes"], 2, 20),
("What does 'Pratikraman' literally mean?", ["Going to temple","Turning back from sins","Fasting","Silent meditation"], 2, 20),
("Which sutra establishes the Guru's presence at the start?", ["Panchindiya","Logassa","Annath","Vandittu"], 1, 20),
("The Khamasamana bow touches how many limbs to the ground?", ["Three","Four","Five","Seven"], 3, 20),
("Tikhutto is recited while doing what?", ["Fasting","Circling the Guru 3 times","Lighting the lamp","Reading scripture"], 2, 20),
("Which sutra is the shravak's confession of the 12 vows?", ["Vandittu","Namutthunam","Jagchintamani","Tassa Uttari"], 1, 20),
("How many verses does the full Vandittu Sutra have?", ["25","36","50","99"], 3, 20),
("Panchindiya describes how many qualities of a Guru?", ["12","24","36","108"], 3, 20),
("'Tassa Michhami Dukkadam' asks for what?", ["Wealth","Forgiveness for wrongs","Good health","Knowledge"], 2, 20),
("Which Pratikraman is done on Samvatsari?", ["Devasi","Raisi","Pakkhi","Samvatsari Pratikraman"], 4, 20),
("Iriyavahiyam repents harm caused while doing what?", ["Speaking","Walking","Eating","Sleeping"], 2, 20),
("Siddhanam Buddhanam says one namaskar to Mahavir can do what?", ["Grant wealth","Cross the ocean of sansara","Cure illness","Win debates"], 2, 20),
]),
"kahoot-4-festivals-practices": ("Jain Festivals & Practices", [
("Paryushan (Shvetambar) lasts how many days?", ["5","8","10","15"], 2, 20),
("The last day of Paryushan is called what?", ["Diwali","Samvatsari","Gyan Panchami","Akshaya Tritiya"], 2, 20),
("What is said to all on Samvatsari?", ["Jai Jinendra","Michami Dukkadam","Uttam Kshama","Namo Arihantanam"], 2, 20),
("Das Lakshan Parva is celebrated by which tradition?", ["Shvetambar","Digambar","Both equally","Neither"], 2, 20),
("For Jains, Diwali marks what event?", ["Rama's return","Mahavir's nirvana","Adinath's birth","New harvest"], 2, 20),
("Which scripture is read during Paryushan?", ["Kalpa Sutra","Tattvartha Sutra","Bhagavati Sutra","Uttaradhyayan"], 1, 20),
("Ayambil Oli is observed how many times a year?", ["Once","Twice","Four times","Monthly"], 2, 20),
("Navkarsi means not eating until when?", ["Sunrise","48 min after sunrise","Noon","Sunset"], 2, 20),
("Chauvihar means no food or water after what?", ["Noon","3 PM","Sunset","Midnight"], 3, 20),
("Akshaya Tritiya celebrates whose fast-breaking?", ["Mahavir","Rishabhdev (Adinath)","Parshvanath","Gautam Swami"], 2, 20),
("Gyan Panchami is dedicated to what?", ["Wealth","Knowledge & scriptures","Forgiveness","Fasting"], 2, 20),
("Ahimsa means what?", ["Truthfulness","Non-violence","Non-attachment","Celibacy"], 2, 20),
]),
"kahoot-5-stavans-famous-lines": ("Stavans & Famous Lines", [
("'Meri Bhavna' is a famous composition about what?", ["Pilgrimage","Noble wishes for all beings","Temple building","Fasting rules"], 2, 20),
("The 'Mangal Divo' is sung while doing what?", ["Fasting","Offering the lamp (aarti)","Walking to temple","Morning bath"], 2, 20),
("'Barah Bhavana' refers to how many reflections?", ["8","10","12","16"], 3, 20),
("'Maitri Bhav' asks for friendship with whom?", ["Family only","Fellow Jains","All living beings","Monks only"], 3, 20),
("How many auspicious dreams did Mahavir's mother see (Shvetambar)?", ["8","12","14","16"], 3, 20),
("'Kshamavani' songs are about what?", ["Wealth","Forgiveness","Harvest","Marriage"], 2, 20),
("A 'stavan' is best described as what?", ["A devotional song","A fasting rule","A temple","A scripture"], 1, 20),
("'Jai Bolo Mahavir' celebrates whom?", ["The 23rd Tirthankara","The 24th Tirthankara","A king","A temple"], 2, 20),
("The 'Navkar Mahamantra' stavan glorifies which prayer?", ["Logassa","Navkar Mantra","Vandittu","Bhaktamar"], 2, 20),
("'Dekho Paryushan Aaya' welcomes what?", ["New year","The festival of Paryushan","Monsoon","A wedding"], 2, 20),
("'Guru Vandana' is a song honoring whom?", ["Parents","Spiritual teachers","Kings","Children"], 2, 20),
("'Swadhyay' means what?", ["Singing loudly","Self-study of scriptures","Temple cleaning","Donating food"], 2, 20),
]),
}

HEADERS = ["Question - max 120 characters",
           "Answer 1 - max 75 characters",
           "Answer 2 - max 75 characters",
           "Answer 3 - max 75 characters",
           "Answer 4 - max 75 characters",
           "Time limit (sec) - 5, 10, 20, 30, 60, 90, 120, or 240 secs",
           "Correct answer(s) - choose at least one"]

os.makedirs("kahoots", exist_ok=True)
for fname, (title, qs) in QUIZZES.items():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws["B2"] = f"Kahoot quiz: {title}"
    ws["B2"].font = Font(name="Arial", bold=True, size=14)
    ws["B3"] = "Import this file at create.kahoot.it - Add question - Import spreadsheet"
    ws["B3"].font = Font(name="Arial", italic=True, size=10)
    ws["A8"] = "Question #"
    for col, h in enumerate(HEADERS, start=2):
        c = ws.cell(row=8, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = PatternFill("solid", start_color="46178F")
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws["A8"].font = Font(name="Arial", bold=True, size=10)
    for i, (q, opts, correct, t) in enumerate(qs, start=1):
        assert len(q) <= 120, f"Q too long: {q}"
        for o in opts: assert len(o) <= 75, f"A too long: {o}"
        r = 8 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=q)
        for j, o in enumerate(opts):
            ws.cell(row=r, column=3 + j, value=o)
        ws.cell(row=r, column=7, value=t)
        ws.cell(row=r, column=8, value=correct)
        for col in range(1, 9):
            ws.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 62
    for col in "CDEF": ws.column_dimensions[col].width = 30
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    wb.save(f"kahoots/{fname}.xlsx")
    print(f"OK {fname}.xlsx ({len(qs)} questions)")
